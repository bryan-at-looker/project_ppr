#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sun Feb  4 17:55:16 2018

@author: Looker
"""


from flask import Flask, request, abort
import openpyxl
#import pandas
import base64
import json
from io import BytesIO
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
                

app = Flask(__name__)


@app.route('/webhook', methods=['POST'])
def webhook():
    if request.method == 'POST':
        medium = Side(border_style="medium", color="8a98ad")
        double = Side(border_style="double", color="ff0000")
        accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        brdr2 = Border(top=medium, left=medium, right=medium, bottom=medium)
        font = Font(b=True, color="FF0000")
        al = Alignment(horizontal="center", vertical="center")
        
        def checkValues(value1,value2):
            return value1 == value2
        
        bg_rgb = ['68BD45', '329bd6','ffffe0']
        bg_rgb_len = len(bg_rgb)
        bg_argb = ["FF" + str(x) for x in bg_rgb]
        
        wh = json.load(BytesIO(request.json))
        data = wh['attachment']['data']
        
        wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(data)))
        # OPEN WORKBOOK IN MEMORY BY CALLING THE ATTACHMENT>>DATA AND DECODING
        ws = wb.active
        
        rows = ws.max_row
        column = ws.max_column
        first_data_row = {"index":2, "cell":3} # index, cell
        first_data_col = {"index":3, "cell":'D'}
        last_data_row = {'index':68,'cell':69}
        last_data_col = {'index':74,'cell':'BW'}
        
        cells = ws['A']
            
        color_counter = 0
        breaker = [first_data_row['cell']]
        
        for i in range(first_data_row['index'],rows):
            #cells[i].fill = openpyxl.styles.PatternFill(start_color=bg_argb[color_counter], end_color=bg_argb[color_counter], fill_type='solid')
            #cells[i].border = medium_sides
            try:
                if not checkValues(cells[i].value, cells[i+1].value):
                    #color_counter = (color_counter+1) % bg_rgb_len
                    breaker.append(i+1+1) # add 1 for the next cell add 1 for index
            except:
                print('last row')
        
        breaker.append(i+1+1) # add the last row + 1
        
        
        for i in range(len(breaker)-1):
            breaker_start = "A" + str(breaker[i])
            breaker_end =   "A" + str(breaker[i+1]-1)
            rng = breaker_start+':'+breaker_end
            #ws.merge_cells(rng)
            ws[breaker_start].alignment = openpyxl.styles.Alignment(vertical='center')
            style_range(ws, rng, border=brdr2
                        , fill=PatternFill("solid", bg_argb[color_counter])
        #                , font=Font(b=True, color="000000")
                        , alignment=Alignment(horizontal="center", vertical="center"))
               
        
        # MANIPULATE COLUMN 2
        
        cells = ws['B']
            
        color_counter = 1
        breaker = [first_data_row['cell']]
        
        for i in range(first_data_row['index'],rows):
            cells[i].fill = openpyxl.styles.PatternFill(start_color=bg_argb[color_counter], end_color=bg_argb[color_counter], fill_type='solid')
            #cells[i].border = medium_sides
            try:
                if not checkValues(cells[i].value, cells[i+1].value):
                    #color_counter = (color_counter+1) % bg_rgb_len
                    breaker.append(i+1+1) # add 1 for the next cell add 1 for index
            except:
                print('last row')
        
        breaker.append(i+1+1) # add the last row + 1
        
        
        for i in range(len(breaker)-1):
            breaker_start = "B" + str(breaker[i])
            breaker_end =   "B" + str(breaker[i+1]-1)
            rng = breaker_start+':'+breaker_end
            #ws.merge_cells(rng)
            ws[breaker_start].alignment = openpyxl.styles.Alignment(vertical='center')
            style_range(ws, rng, border=brdr2
                        , fill=PatternFill("solid", bg_argb[color_counter])
                        , font=Font(b=True, color="000000")
                        , alignment=Alignment(horizontal="center", vertical="center", text_rotation=90))
            
            
        # MANIPULATE FIRST ROW
        cells = ws[1]
            
        color_counter = 2
        breaker = [first_data_col['cell']]
        
        for i in range(first_data_col['index'],column):
            cells[i].fill = openpyxl.styles.PatternFill(start_color=bg_argb[color_counter], end_color=bg_argb[color_counter], fill_type='solid')
            #cells[i].border = medium_sides
            try:
                if not checkValues(cells[i].value, cells[i+1].value):
                    #color_counter = (color_counter+1) % bg_rgb_len
                    breaker.append(cells[i+1].column) # add 1 for the next cell add 1 for index
            except:
                print('last row')
        
        breaker.append(ws.cell(row=1, column = i+1+1).column) # add the last row + 1
        
        
        for i in range(1,len(breaker)):
            breaker_start = breaker[i-1] + str(1)
            breaker_end =   ws.cell(row=1,column=openpyxl.utils.column_index_from_string(breaker[i])-1).column + str(1)
            rng = breaker_start+':'+breaker_end
            #ws.merge_cells(rng)
        #    ws[breaker_start].alignment = openpyxl.styles.Alignment(vertical='center')
            style_range(ws, rng, border=brdr2
                        , fill=PatternFill("solid", bg_argb[color_counter])
                        , font=Font(b=True, color="000000")
                        , alignment=Alignment(horizontal="center", vertical="center"))
        
        # MANIPULATE SECOND ROW
        cells = ws[2]
        
        total_column = []
        
        for i in range(first_data_col['index'],column):
            cells[i].font = openpyxl.styles.Font(italic=True, bold=True)
            cells[i].fill = PatternFill("solid", bg_argb[color_counter])
            cells[i].alignment = Alignment(horizontal="center")
            cells[i].border = brdr2
            if cells[i].value == 'Total':
                total_column.append(i)
        
        #MANIPUTATE TOTALS
        
        for i in range(len(total_column)):
            for j in range(first_data_row['cell'],last_data_row['cell']+1):
                ws.cell(column=total_column[i]+1, row=j).fill = PatternFill("solid", 'ffffe0')
                ws.cell(column=total_column[i]+1, row=j).border = Border(left=medium, right=medium)
        
            
        wb.save('output_test2.xlsx')


        print(request.json)
        return '', 200
    else:
        abort(400)


if __name__ == '__main__':
    app.run(host= '0.0.0.0')